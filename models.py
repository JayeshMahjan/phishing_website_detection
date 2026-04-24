"""
PhishGuard AI - Excel-based Data Storage
Stores user data and URL analyses in Excel files instead of database.
Uses openpyxl for Excel file manipulation.
"""

import openpyxl
from openpyxl import Workbook
import os
import bcrypt
from datetime import datetime
import json


class UserStorage:

    """Manage user accounts in Excel file.
    This class manages user accounts, registration, and authentication data."""
    
    def __init__(self, filename='data/users.xlsx'):
        self.filename = filename
        self._ensure_file_exists()
    
    def _ensure_file_exists(self):
        """Create Excel file with headers if it doesn't exist."""
        os.makedirs(os.path.dirname(self.filename), exist_ok=True)
        
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"
            
            # Add headers
            headers = ['username', 'email', 'password_hash', 'created_at', 'last_login']
            ws.append(headers)
            
            wb.save(self.filename)
    
    def _load_workbook(self):
        """Load the workbook."""
        return openpyxl.load_workbook(self.filename)
    
    def find_by_username(self, username):
        """Find user by username."""
        wb = self._load_workbook()
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username:  # username column
                return {
                    'username': row[0],
                    'email': row[1],
                    'password_hash': row[2],
                    'created_at': row[3],
                    'last_login': row[4]
                }
        return None
    
    def find_by_email(self, email):
        """Find user by email."""
        wb = self._load_workbook()
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == email:  # email column
                return {
                    'username': row[0],
                    'email': row[1],
                    'password_hash': row[2],
                    'created_at': row[3],
                    'last_login': row[4]
                }
        return None
    
    def find_by_username_or_email(self, username_or_email):
        """Find user by either username or email."""
        user = self.find_by_username(username_or_email)
        if not user:
            user = self.find_by_email(username_or_email)
        return user
    
    def get_all_users(self):
        """Get all users from Excel file."""
        wb = self._load_workbook()
        ws = wb.active
        
        all_users = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If username exists
                all_users.append({
                    'username': row[0],
                    'email': row[1],
                    'password_hash': row[2],
                    'created_at': row[3],
                    'last_login': row[4]
                })
        
        return all_users
    
    def verify_password(self, user, password):
        """Verify password against stored hash."""
        if not user or not user.get('password_hash'):
            return False
        
        try:
            return bcrypt.checkpw(
                password.encode('utf-8'),
                user['password_hash'].encode('utf-8')
            )
        except Exception:
            return False
    
    def create_user(self, username, email, password):
        """Create a new user and save to Excel."""
        wb = self._load_workbook()
        ws = wb.active
        
        # Hash password
        password_hash = bcrypt.hashpw(
            password.encode('utf-8'),
            bcrypt.gensalt(rounds=12)
        ).decode('utf-8')
        
        # Get current timestamp
        created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Add new user
        ws.append([username, email, password_hash, created_at, ''])
        
        # Save workbook
        wb.save(self.filename)
        
        return {
            'username': username,
            'email': email,
            'created_at': created_at
        }
    
    def update_last_login(self, username):
        """Update last login timestamp for user."""
        wb = self._load_workbook()
        ws = wb.active
        
        last_login = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Find and update user
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=5, values_only=False):
            if row[0].value == username:  # username column
                row[4].value = last_login  # last_login column
                break
        
        wb.save(self.filename)


class URLAnalysisStorage:
    """Manage URL analysis history in Excel file."""
    
    def __init__(self, filename='data/url_analyses.xlsx'):
        self.filename = filename
        self._ensure_file_exists()
    
    def _ensure_file_exists(self):
        """Create Excel file with headers if it doesn't exist."""
        os.makedirs(os.path.dirname(self.filename), exist_ok=True)
        
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Analyses"
            
            # Add headers
            headers = [
                'id', 'username', 'url', 'prediction', 'confidence',
                'legitimacy_score', 'phishing_score', 'features_json',
                'analyzed_at', 'ip_address', 'user_agent'
            ]
            ws.append(headers)
            
            wb.save(self.filename)
    
    def _load_workbook(self):
        """Load the workbook."""
        return openpyxl.load_workbook(self.filename)
    
    def _get_next_id(self):
        """Get next available ID."""
        wb = self._load_workbook()
        ws = wb.active
        
        max_id = 0
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
            if row[0] and row[0] > max_id:
                max_id = row[0]
        
        return max_id + 1
    
    def add_analysis(self, username, url, prediction, confidence,
                     legitimacy_score, phishing_score, features,
                     ip_address='', user_agent=''):
        """Add a new URL analysis to Excel."""
        wb = self._load_workbook()
        ws = wb.active
        
        # Get next ID
        analysis_id = self._get_next_id()
        
        # Get current timestamp
        analyzed_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Convert features to JSON string
        features_json = json.dumps(features)
        
        # Add new analysis
        ws.append([
            analysis_id,
            username,
            url,
            prediction,
            confidence,
            legitimacy_score,
            phishing_score,
            features_json,
            analyzed_at,
            ip_address,
            user_agent
        ])
        
        # Save workbook
        wb.save(self.filename)
        
        return analysis_id
    
    def get_user_analyses(self, username, limit=10):
        """Get analyses for a specific user (most recent first)."""
        wb = self._load_workbook()
        ws = wb.active
        
        user_analyses = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == username:  # username column
                user_analyses.append({
                    'id': row[0],
                    'url': row[2],
                    'prediction': row[3],
                    'confidence': row[4],
                    'legitimacy_score': row[5],
                    'phishing_score': row[6],
                    'analyzed_at': row[8]
                })
        
        # Sort by date (most recent first) and limit
        user_analyses.sort(key=lambda x: x['analyzed_at'], reverse=True)
        return user_analyses[:limit]
    
    def get_all_analyses(self):
        """Get all analyses."""
        wb = self._load_workbook()
        ws = wb.active
        
        all_analyses = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            all_analyses.append({
                'id': row[0],
                'username': row[1],
                'url': row[2],
                'prediction': row[3],
                'confidence': row[4],
                'analyzed_at': row[8]
            })
        
        return all_analyses
